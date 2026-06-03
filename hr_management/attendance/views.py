from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db import transaction, models
from django.db.models import Prefetch
from decimal import Decimal
import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from .models import (WorkShift, AttendanceRecord, DailyAttendance, 
                     AttendanceSummary, EmployeeAttendance)
from .forms import DailyAttendanceForm
from .utils import get_attendance_summary_data
from employees.models import Employee, Position
from payroll.models import Payroll, PayrollDetail


def dashboard(request):
    return redirect('attendance:daily_attendance_form')

def work_shift_list(request):
    work_shifts = WorkShift.objects.all().order_by('code')
    paginator = Paginator(work_shifts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'attendance/work_shift_list.html', {
        'work_shifts': page_obj,
        'page_obj': page_obj,
    })

def work_shift_form(request, id=None):
    if id:
        work_shift = get_object_or_404(WorkShift, id=id)
    else:
        work_shift = None

    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        start_time = request.POST.get('start_time')
        check_in_start = request.POST.get('check_in_start')
        check_in_end = request.POST.get('check_in_end')
        end_time = request.POST.get('end_time')
        check_out_start = request.POST.get('check_out_start')
        check_out_end = request.POST.get('check_out_end')
        has_break = request.POST.get('has_break') == 'on'
        work_hours = float(request.POST.get('work_hours', 0))
        work_days = float(request.POST.get('work_days', 1))
        normal_day_coefficient = float(request.POST.get('normal_day_coefficient', 1))
        rest_day_coefficient = float(request.POST.get('rest_day_coefficient', 2))
        holiday_coefficient = float(request.POST.get('holiday_coefficient', 3))
        deduct_if_no_check_in = request.POST.get('deduct_if_no_check_in') == 'on'
        deduct_if_no_check_out = request.POST.get('deduct_if_no_check_out') == 'on'
        apply_to_all = request.POST.get('apply_to_all') == 'yes'
        employee_ids = request.POST.getlist('employees')

        if work_shift:  # Sửa ca làm việc
            work_shift.name = name
            work_shift.code = code
            work_shift.start_time = start_time
            work_shift.check_in_start = check_in_start
            work_shift.check_in_end = check_in_end
            work_shift.end_time = end_time
            work_shift.check_out_start = check_out_start
            work_shift.check_out_end = check_out_end
            work_shift.has_break = has_break
            work_shift.work_hours = work_hours
            work_shift.work_days = work_days
            work_shift.normal_day_coefficient = normal_day_coefficient
            work_shift.rest_day_coefficient = rest_day_coefficient
            work_shift.holiday_coefficient = holiday_coefficient
            work_shift.deduct_if_no_check_in = deduct_if_no_check_in
            work_shift.deduct_if_no_check_out = deduct_if_no_check_out
            work_shift.apply_to_all = apply_to_all
            work_shift.save()
            if Employee:
                work_shift.employees.set(employee_ids)
            messages.success(request, "Cập nhật ca làm việc thành công!")
        else:  # Thêm ca làm việc mới
            work_shift = WorkShift.objects.create(
                name=name,
                code=code,
                start_time=start_time,
                check_in_start=check_in_start,
                check_in_end=check_in_end,
                end_time=end_time,
                check_out_start=check_out_start,
                check_out_end=check_out_end,
                has_break=has_break,
                work_hours=work_hours,
                work_days=work_days,
                normal_day_coefficient=normal_day_coefficient,
                rest_day_coefficient=rest_day_coefficient,
                holiday_coefficient=holiday_coefficient,
                deduct_if_no_check_in=deduct_if_no_check_in,
                deduct_if_no_check_out=deduct_if_no_check_out,
                apply_to_all=apply_to_all
            )
            if Employee and employee_ids:
                work_shift.employees.set(employee_ids)
            messages.success(request, "Thêm ca làm việc thành công!")

        return redirect('attendance:work_shift_list')  # Sửa ở đây

    employees = Employee.objects.all() if Employee else []
    return render(request, 'attendance/work_shift_form.html', {
        'work_shift': work_shift,
        'employees': employees
    })

def work_shift_detail(request, id):
    work_shift = get_object_or_404(WorkShift, id=id)
    return render(request, 'attendance/work_shift_detail.html', {'work_shift': work_shift})

def work_shift_delete(request, id):
    work_shift = get_object_or_404(WorkShift, id=id)
    work_shift.delete()
    messages.success(request, f"Đã xóa ca làm việc {work_shift.name} thành công!")
    return redirect('attendance:work_shift_list')  # Thêm namespace 'attendance'

def attendance_detail_list(request):
    attendance_records = AttendanceRecord.objects.all()
    work_shifts = WorkShift.objects.all()

    # Thêm phân trang
    paginator = Paginator(attendance_records, 10)  # 10 bản ghi mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'attendance/attendance_detail_list.html', {
        'attendance_records': page_obj,
        'work_shifts': work_shifts,
        'page_obj': page_obj
    })

def attendance_detail_save(request):
    if request.method == 'POST':
        # Get necessary data from POST request
        attendance_record_id = request.POST.get('attendance_record_id')
        employee_id = request.POST.get('employee_id')
        check_in_time = request.POST.get('check_in_time')
        check_out_time = request.POST.get('check_out_time')

        # Handle saving or updating the attendance record here
        attendance_record = AttendanceRecord.objects.get(id=attendance_record_id)
        # Update attendance record fields as needed
        # Example: attendance_record.check_in_time = check_in_time
        attendance_record.save()

        messages.success(request, "Attendance record saved successfully!")
        return redirect('attendance_detail_list')  # Redirect to the attendance detail list or another page

    return redirect('attendance_detail_list')  # Default redirect if method is not POST

def attendance_detail_form(request, id=None):
    if id:
        attendance_record = get_object_or_404(AttendanceRecord, id=id)
    else:
        attendance_record = None

    if request.method == 'POST':
        name = request.POST.get('name')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        attendance_type = request.POST.get('attendance_type')
        position_id = request.POST.get('positions')
        apply_to_all_shifts = request.POST.get('apply_to_all_shifts') == 'on'
        work_shift_ids = request.POST.getlist('work_shifts')

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

        position = get_object_or_404(Position, id=position_id)

        if attendance_record:  # Sửa bảng chấm công chi tiết
            attendance_record.name = name
            attendance_record.start_date = start_date
            attendance_record.end_date = end_date
            attendance_record.attendance_type = attendance_type
            attendance_record.positions = position
            attendance_record.apply_to_all_shifts = apply_to_all_shifts
            attendance_record.save()
            if not apply_to_all_shifts:
                attendance_record.work_shifts.set(work_shift_ids)
            else:
                attendance_record.work_shifts.clear()
            messages.success(request, "Cập nhật bảng chấm công chi tiết thành công!")
        else:  # Thêm mới bảng chấm công chi tiết
            attendance_record = AttendanceRecord.objects.create(
                name=name,
                start_date=start_date,
                end_date=end_date,
                attendance_type=attendance_type,
                positions=position,
                apply_to_all_shifts=apply_to_all_shifts
            )
            if not apply_to_all_shifts:
                attendance_record.work_shifts.set(work_shift_ids)
            messages.success(request, "Thêm bảng chấm công chi tiết thành công!")

        return redirect('attendance:attendance_detail_list')  # Sửa ở đây

    positions = Position.objects.all()
    work_shifts = WorkShift.objects.all()
    return render(request, 'attendance/attendance_detail_form.html', {
        'attendance_record': attendance_record,
        'positions': positions,
        'work_shifts': work_shifts
    })

def attendance_detail_delete(request, id):
    attendance_record = get_object_or_404(AttendanceRecord, id=id)
    attendance_record.delete()
    messages.success(request, "Đã xóa bảng chấm công chi tiết thành công!")
    return redirect('attendance:attendance_detail_list')  # Thêm namespace 'attendance'

@login_required
def attendance_detail_view(request, id):
    attendance_record = get_object_or_404(AttendanceRecord, id=id)
    employees = Employee.objects.filter(position=attendance_record.positions).distinct()

    # Tạo danh sách ngày (bao gồm cả Chủ nhật)
    date_list = []
    current_date = attendance_record.start_date
    while current_date <= attendance_record.end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)

    attendance_data = {}
    for employee in employees:
        attendance_data[employee.id] = {}
        for date in date_list:
            daily_attendance = DailyAttendance.objects.filter(
                attendance_record=attendance_record,
                employee=employee,
                date=date
            ).first()
            is_on_time = False
            is_enough_work = False
            if daily_attendance and daily_attendance.check_in_time and daily_attendance.work_shift:
                check_in_time = daily_attendance.check_in_time
                shift_start = daily_attendance.work_shift.check_in_end
                is_on_time = check_in_time <= shift_start
                if daily_attendance.check_out_time:
                    work_duration = (
                        daily_attendance.check_out_time.hour * 60 + daily_attendance.check_out_time.minute) - \
                        (daily_attendance.check_in_time.hour * 60 + daily_attendance.check_in_time.minute)
                    required_duration = (
                        daily_attendance.work_shift.end_time.hour * 60 + daily_attendance.work_shift.end_time.minute) - \
                        (daily_attendance.work_shift.start_time.hour * 60 + daily_attendance.work_shift.start_time.minute)
                    is_enough_work = work_duration >= required_duration * 0.9
            attendance_data[employee.id][date] = {
                'attendance': daily_attendance,
                'is_on_time': is_on_time,
                'is_enough_work': is_enough_work
            }

    context = {
        'attendance_record': attendance_record,
        'employees': employees,
        'date_list': date_list,
        'attendance_data': attendance_data,
    }
    return render(request, 'attendance/attendance_detail_view.html', context)

@login_required
def export_attendance_to_excel(request, record_id):
    """Xuất dữ liệu chấm công chi tiết ra file Excel"""
    attendance_record = get_object_or_404(AttendanceRecord, id=record_id)
    employees = Employee.objects.filter(position=attendance_record.positions).distinct()

    # Tạo danh sách ngày (trừ Chủ nhật)
    date_list = []
    current_date = attendance_record.start_date
    while current_date <= attendance_record.end_date:
        if current_date.weekday() != 6:  # Bỏ qua Chủ nhật
            date_list.append(current_date)
        current_date += timedelta(days=1)

    # Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang Cham Cong Chi Tiet"

    # Tiêu đề bảng
    ws['A1'] = "Nhân viên"
    for idx, date in enumerate(date_list, start=2):
        col_letter = get_column_letter(idx)
        ws[f'{col_letter}1'] = date.strftime('%a/%d')

    # Dữ liệu
    for row_idx, employee in enumerate(employees, start=2):
        ws[f'A{row_idx}'] = f"{employee.last_name} {employee.first_name} ({employee.code})"
        for col_idx, date in enumerate(date_list, start=2):
            col_letter = get_column_letter(col_idx)
            daily_attendance = DailyAttendance.objects.filter(
                attendance_record=attendance_record,
                employee=employee,
                date=date
            ).first()
            if daily_attendance:
                if daily_attendance.check_in_time:
                    check_in = daily_attendance.check_in_time.strftime('%H:%M')
                    check_out = daily_attendance.check_out_time.strftime(
                        '%H:%M') if daily_attendance.check_out_time else ''
                    status = f"{check_in} - {check_out}"
                else:
                    status = {
                        'permitted_absence': 'Nghỉ có phép',
                        'unpermitted_absence': 'Nghỉ không phép',
                        'regime_absence': 'Nghỉ chế độ',
                        'not_absent': '--:--'
                    }.get(daily_attendance.attendance_status, '--:--')
            else:
                status = '--:--'
            ws[f'{col_letter}{row_idx}'] = status

    # Định dạng bảng
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Tạo response để tải file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bang_cham_cong_chi_tiet_{record_id}.xlsx"'
    wb.save(response)
    return response

@login_required
def import_attendance_from_excel(request, record_id):
    """Nhập dữ liệu chấm công chi tiết từ file Excel"""
    attendance_record = get_object_or_404(AttendanceRecord, id=record_id)

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Vui lòng chọn file Excel để tải lên!")
            return redirect('attendance_detail_view', id=record_id)

        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "File phải có định dạng .xlsx!")
            return redirect('attendance_detail_view', id=record_id)

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Lấy danh sách ngày từ tiêu đề
            date_list = []
            for col in range(2, ws.max_column + 1):
                date_str = ws[f'{get_column_letter(col)}1'].value
                if date_str:
                    try:
                        date_obj = datetime.strptime(date_str, '%a/%d').replace(year=attendance_record.start_date.year)
                        date_list.append(date_obj.date())
                    except ValueError:
                        continue

            # Xử lý từng dòng dữ liệu
            for row in range(2, ws.max_row + 1):
                employee_cell = ws[f'A{row}'].value
                if not employee_cell:
                    continue

                # Tìm nhân viên từ mã (giả sử mã nằm trong ngoặc)
                try:
                    employee_code = employee_cell.split('(')[-1].strip(')')
                    employee = Employee.objects.get(code=employee_code)
                except Employee.DoesNotExist:
                    messages.warning(request, f"Không tìm thấy nhân viên với mã {employee_code}, bỏ qua dòng này.")
                    continue

                # Xử lý dữ liệu chấm công cho từng ngày
                for col, date in enumerate(date_list, start=2):
                    cell_value = ws[f'{get_column_letter(col)}{row}'].value
                    if not cell_value or cell_value == '--:--':
                        continue

                    # Tìm hoặc tạo bản ghi DailyAttendance
                    daily_attendance, created = DailyAttendance.objects.get_or_create(
                        attendance_record=attendance_record,
                        employee=employee,
                        date=date,
                        defaults={'attendance_status': 'not_absent'}
                    )

                    # Xử lý dữ liệu từ ô Excel
                    if 'Nghỉ' in cell_value:
                        if 'có phép' in cell_value:
                            daily_attendance.attendance_status = 'permitted_absence'
                        elif 'không phép' in cell_value:
                            daily_attendance.attendance_status = 'unpermitted_absence'
                        elif 'chế độ' in cell_value:
                            daily_attendance.attendance_status = 'regime_absence'
                        daily_attendance.check_in_time = None
                        daily_attendance.check_out_time = None
                    else:
                        try:
                            times = cell_value.split(' - ')
                            if len(times) == 2:
                                check_in = datetime.strptime(times[0], '%H:%M').time()
                                check_out = datetime.strptime(times[1], '%H:%M').time() if times[1] else None
                                daily_attendance.check_in_time = check_in
                                daily_attendance.check_out_time = check_out
                                daily_attendance.attendance_status = 'not_absent'
                        except ValueError:
                            messages.warning(request,
                                             f"Dữ liệu không hợp lệ tại ô {get_column_letter(col)}{row}, bỏ qua.")
                            continue

                    # Gán ca làm việc
                    work_shift = WorkShift.objects.filter(
                        models.Q(employees=employee) | models.Q(apply_to_all=True)
                    ).first()
                    daily_attendance.work_shift = work_shift
                    daily_attendance.save()

            messages.success(request, "Đã nhập dữ liệu chấm công từ file Excel thành công!")
        except Exception as e:
            messages.error(request, f"Có lỗi khi nhập file Excel: {str(e)}")

        return redirect('attendance_detail_view', id=record_id)

    return redirect('attendance_detail_view', id=record_id)

@login_required
def get_attendance_detail(request, record_id, employee_id, date):
    """Lấy thông tin chấm công chi tiết của một nhân viên vào một ngày cụ thể"""
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        attendance_record = get_object_or_404(AttendanceRecord, id=record_id)
        employee = get_object_or_404(Employee, id=employee_id)

        daily_attendance = DailyAttendance.objects.filter(
            attendance_record=attendance_record,
            employee=employee,
            date=date_obj
        ).first()

        data = {
            'paid_work_days': daily_attendance.paid_work_days if daily_attendance else 1,
            'actual_work_days': daily_attendance.actual_work_days if daily_attendance else 1,
            'check_in_time': daily_attendance.check_in_time.strftime(
                '%H:%M') if daily_attendance and daily_attendance.check_in_time else '',
            'check_out_time': daily_attendance.check_out_time.strftime(
                '%H:%M') if daily_attendance and daily_attendance.check_out_time else '',
            'attendance_status': daily_attendance.attendance_status if daily_attendance else 'not_absent',
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def update_attendance_detail(request, record_id, employee_id, date):
    """Cập nhật thông tin chấm công chi tiết của một nhân viên vào một ngày cụ thể"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không hợp lệ!'}, status=405)

    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        attendance_record = get_object_or_404(AttendanceRecord, id=record_id)
        employee = get_object_or_404(Employee, id=employee_id)

        # Lấy dữ liệu từ request body
        data = json.loads(request.body)
        paid_work_days = float(data.get('paid_work_days', 1))
        actual_work_days = float(data.get('actual_work_days', 1))
        check_in_time = data.get('check_in_time')
        check_out_time = data.get('check_out_time')
        attendance_status = data.get('attendance_status', 'not_absent')

        # Tìm hoặc tạo bản ghi DailyAttendance
        daily_attendance, created = DailyAttendance.objects.get_or_create(
            attendance_record=attendance_record,
            employee=employee,
            date=date_obj,
            defaults={'attendance_status': 'not_absent'}
        )

        # Cập nhật dữ liệu
        daily_attendance.paid_work_days = paid_work_days
        daily_attendance.actual_work_days = actual_work_days
        daily_attendance.attendance_status = attendance_status

        if attendance_status == 'not_absent' and check_in_time:
            daily_attendance.check_in_time = datetime.strptime(check_in_time, '%H:%M').time()
            daily_attendance.check_out_time = datetime.strptime(check_out_time,
                                                                '%H:%M').time() if check_out_time else None
        else:
            daily_attendance.check_in_time = None
            daily_attendance.check_out_time = None

        # Gán ca làm việc
        work_shift = WorkShift.objects.filter(
            models.Q(employees=employee) | models.Q(apply_to_all=True)
        ).first()
        daily_attendance.work_shift = work_shift
        daily_attendance.save()

        return JsonResponse({'status': 'success', 'message': 'Lưu dữ liệu thành công!'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def update_daily_attendance(request, record_id, employee_id, date_str):
    if request.method == 'POST':
        try:
            # Lấy dữ liệu JSON từ request.body
            data = json.loads(request.body)
            print('Dữ liệu nhận được từ client:', data)  # Debug dữ liệu nhận được

            attendance_record = get_object_or_404(AttendanceRecord, id=record_id)
            employee = get_object_or_404(Employee, id=employee_id)
            date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Lấy hoặc tạo bản ghi chấm công hàng ngày
            daily_attendance, created = DailyAttendance.objects.get_or_create(
                attendance_record=attendance_record,
                employee=employee,
                date=date,
                defaults={'attendance_status': 'not_absent'}  # Mặc định là "Không nghỉ" nếu chưa chấm
            )

            # Lấy dữ liệu từ payload
            paid_work_days = float(data.get('paid_work_days', 1))
            actual_work_days = float(data.get('actual_work_days', 1))
            check_in_time_str = data.get('check_in_time') or None
            check_out_time_str = data.get('check_out_time') or None
            attendance_status = data.get('attendance_status')

            # Kiểm tra attendance_status
            valid_statuses = [choice[0] for choice in DailyAttendance.ATTENDANCE_STATUS_CHOICES]
            if not attendance_status or attendance_status not in valid_statuses:
                return JsonResponse({'status': 'error', 'message': f'Trạng thái nghỉ không hợp lệ! Nhận được: "{attendance_status}", giá trị hợp lệ: {valid_statuses}'})

            # Chuyển đổi thời gian từ chuỗi (HH:MM)
            check_in_time = None
            check_out_time = None
            if check_in_time_str and check_in_time_str != 'null':
                try:
                    check_in_time = datetime.strptime(check_in_time_str, '%H:%M').time()
                except ValueError as e:
                    return JsonResponse({'status': 'error', 'message': f'Định dạng giờ vào không hợp lệ (HH:MM): {check_in_time_str}, lỗi: {str(e)}'})
            if check_out_time_str and check_out_time_str != 'null':
                try:
                    check_out_time = datetime.strptime(check_out_time_str, '%H:%M').time()
                except ValueError as e:
                    return JsonResponse({'status': 'error', 'message': f'Định dạng giờ ra không hợp lệ (HH:MM): {check_out_time_str}, lỗi: {str(e)}'})

            # Gán work_shift từ AttendanceRecord
            # Giả sử chỉ có một ca làm việc được chọn trong AttendanceRecord
            work_shift = None
            if attendance_record.apply_to_all_shifts:
                # Nếu áp dụng tất cả ca, lấy ca phù hợp với employee (cần logic cụ thể hơn)
                work_shift = WorkShift.objects.filter(employees=employee).first()
            else:
                # Lấy ca đầu tiên từ danh sách work_shifts của AttendanceRecord
                work_shift = attendance_record.work_shifts.first()

            if not work_shift:
                return JsonResponse({'status': 'error', 'message': 'Không tìm thấy ca làm việc phù hợp!'})

            # Cập nhật bản ghi
            daily_attendance.paid_work_days = paid_work_days
            daily_attendance.actual_work_days = actual_work_days
            daily_attendance.check_in_time = check_in_time
            daily_attendance.check_out_time = check_out_time
            daily_attendance.attendance_status = attendance_status
            daily_attendance.work_shift = work_shift  # Gán work_shift
            daily_attendance.save()

            return JsonResponse({'status': 'success', 'message': 'Chấm công thành công!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Lỗi khi lưu dữ liệu: {str(e)}'})
    return JsonResponse({'status': 'error', 'message': 'Yêu cầu không hợp lệ!'})

def get_daily_attendance(request, record_id, employee_id, date_str):
    try:
        attendance_record = get_object_or_404(AttendanceRecord, id=record_id)
        employee = get_object_or_404(Employee, id=employee_id)
        date = datetime.strptime(date_str, '%Y-%m-%d').date()

        daily_attendance = DailyAttendance.objects.filter(
            attendance_record=attendance_record,
            employee=employee,
            date=date
        ).first()

        if daily_attendance:
            data = {
                'paid_work_days': daily_attendance.paid_work_days,
                'actual_work_days': daily_attendance.actual_work_days,
                'check_in_time': daily_attendance.check_in_time.strftime('%H:%M') if daily_attendance.check_in_time else '',
                'check_out_time': daily_attendance.check_out_time.strftime('%H:%M') if daily_attendance.check_out_time else '',
                'attendance_status': daily_attendance.attendance_status,
            }
        else:
            data = {
                'paid_work_days': 1,
                'actual_work_days': 1,
                'check_in_time': '',
                'check_out_time': '',
                'attendance_status': 'not_absent',
            }

        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Lỗi khi lấy dữ liệu: {str(e)}'})

def attendance_summary(request):
    attendance_summaries = AttendanceSummary.objects.all().order_by('-year', '-month', 'name')
    paginator = Paginator(attendance_summaries, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    positions = Position.objects.all()  # Thêm danh sách vị trí
    attendance_records = AttendanceRecord.objects.all()  # Thêm danh sách bảng chấm công chi tiết
    return render(request, 'attendance/attendance_summary.html', {
        'attendance_summaries': page_obj,
        'page_obj': page_obj,
        'positions': positions,
        'attendance_records': attendance_records
    })

def attendance_summary_form(request, id=None):
    if id:
        attendance_summary = get_object_or_404(AttendanceSummary, id=id)
    else:
        attendance_summary = None

    positions = Position.objects.all()
    attendance_records = AttendanceRecord.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name')
        position_id = request.POST.get('position')
        attendance_record_ids = request.POST.getlist('attendance_records')

        position = get_object_or_404(Position, id=position_id)
        attendance_records_selected = AttendanceRecord.objects.filter(id__in=attendance_record_ids)

        if attendance_summary:
            # Sửa bảng chấm công tổng hợp
            attendance_summary.name = name
            attendance_summary.position = position
            attendance_summary.save()
            attendance_summary.attendance_records.set(attendance_records_selected)
            messages.success(request, "Cập nhật bảng chấm công tổng hợp thành công!")
        else:
            # Tính start_date và end_date từ attendance_records
            if attendance_records_selected:
                start_date = min(record.start_date for record in attendance_records_selected)
                end_date = max(record.end_date for record in attendance_records_selected)
            else:
                # Nếu không có attendance_records, đặt giá trị mặc định
                start_date = datetime.now().date()
                end_date = start_date

            # Thêm bảng chấm công tổng hợp
            attendance_summary = AttendanceSummary.objects.create(
                name=name,
                position=position,
                start_date=start_date,
                end_date=end_date,
                month=start_date.month,
                year=start_date.year
            )
            attendance_summary.attendance_records.set(attendance_records_selected)
            messages.success(request, "Thêm bảng chấm công tổng hợp thành công!")

        return redirect('attendance:attendance_summary')

    return render(request, 'attendance/attendance_summary_form.html', {
        'attendance_summary': attendance_summary,
        'positions': positions,
        'attendance_records': attendance_records,
    })

def attendance_summary_edit(request, id):
    summary = get_object_or_404(AttendanceSummary, id=id)
    # Logic để chỉnh sửa tại đây (form xử lý, POST/GET, ...)
    return render(request, 'attendance/attendance_summary_edit.html', {'summary': summary})

def attendance_summary_list(request):
    # Your view logic here
    return render(request, 'attendance/attendance_summary_list.html')

@login_required
def attendance_list(request):
    """Danh sách bảng chấm công tổng hợp"""
    # Lấy danh sách bảng chấm công tổng hợp
    attendance_summaries = AttendanceSummary.objects.all().order_by('-created_at')

    # Phân trang
    paginator = Paginator(attendance_summaries, 10)  # 10 bảng chấm công mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    return render(request, 'attendance/attendance_list.html', context)

@login_required
def attendance_summary_view(request, id):
    """Xem chi tiết bảng chấm công tổng hợp"""
    attendance_summary = get_object_or_404(AttendanceSummary, id=id)

    # Lấy dữ liệu chấm công tổng hợp
    from .utils import get_attendance_summary_data
    employee_data = get_attendance_summary_data(attendance_summary)  # Truyền đối tượng thay vì ID

    context = {
        'attendance_summary': attendance_summary,
        'employee_data': employee_data,
    }

    return render(request, 'attendance/attendance_summary_view.html', context)

def attendance_summary_delete(request, id):
    attendance_summary = get_object_or_404(AttendanceSummary, id=id)
    attendance_summary.delete()
    messages.success(request, "Đã xóa bảng chấm công tổng hợp thành công!")
    return redirect('attendance:attendance_summary')  # Thêm namespace 'attendance'

@login_required
def transfer_to_payroll(request, summary_id):
    """Chuyển dữ liệu từ bảng chấm công tổng hợp sang tính lương"""
    attendance_summary = get_object_or_404(AttendanceSummary, pk=summary_id)

    # Kiểm tra xem bảng chấm công đã được chuyển sang tính lương chưa
    if attendance_summary.transferred:
        messages.warning(request, 'Bảng chấm công này đã được chuyển sang tính lương trước đó')
        return redirect('attendance:attendance_summary_view', id=summary_id)

    # Lấy dữ liệu nhân viên từ hàm helper để hiển thị trong trang xác nhận
    employee_data = get_attendance_summary_data(summary_id)

    # Gán context trước khi sử dụng trong bất kỳ nhánh nào
    context = {
        'attendance_summary': attendance_summary,
        'employee_data': employee_data,
    }

    if request.method == 'POST':
        # Nếu người dùng đã xác nhận, thực hiện chuyển dữ liệu
        if 'confirm' in request.POST:
            try:
                # Xác định month và year từ attendance_summary hoặc sử dụng thời gian hiện tại
                month = getattr(attendance_summary, 'month', datetime.now().month)
                year = getattr(attendance_summary, 'year', datetime.now().year)

                # Tạo bảng lương mới
                payroll_name = f"Bảng lương tháng {month}/{year} - {attendance_summary.position.name}"
                payroll = Payroll.objects.create(
                    user=request.user,
                    name=payroll_name,
                    month=month,
                    year=year,
                    position=attendance_summary.position,
                    status='draft',
                    created_by=request.user
                )

                # Tính lương cho từng nhân viên
                for employee_info in employee_data:
                    employee = employee_info['employee']
                    basic_salary = getattr(employee, 'basic_salary', 0) or 0
                    standard_work_days = employee_info.get('standard_work_days', 0)
                    actual_workdays = employee_info.get('actual_workdays', 0)
                    unpaid_leave = employee_info.get('unpaid_leave', 0)

                    try:
                        attendance_ratio = float(employee_info.get('attendance_ratio', 0))
                        attendance_ratio = max(0, min(1, attendance_ratio))
                    except (TypeError, ValueError):
                        attendance_ratio = 0

                    try:
                        gross_salary = int(basic_salary * Decimal(str(attendance_ratio)))
                    except Exception:
                        gross_salary = 0

                    deductions_amount = int(gross_salary * Decimal('0.1'))
                    net_salary = gross_salary - deductions_amount

                    PayrollDetail.objects.create(
                        payroll=payroll,
                        employee=employee,
                        basic_salary=basic_salary,
                        attendance_ratio=attendance_ratio,
                        standard_workdays=standard_work_days,
                        actual_workdays=actual_workdays,
                        unpaid_leave=unpaid_leave,
                        gross_salary=gross_salary,
                        deduction_amount=deductions_amount,
                        net_salary=net_salary
                    )

                attendance_summary.transferred = True
                attendance_summary.save()

                messages.success(request,
                                 f'Đã chuyển dữ liệu từ bảng chấm công "{attendance_summary.name}" sang bảng lương thành công')

                # Chuyển hướng thay vì render lại template
                return redirect('attendance:attendance_summary_view', id=summary_id)
            except Exception as e:
                messages.error(request, f'Lỗi khi chuyển dữ liệu: {str(e)}')
                return redirect('attendance:attendance_summary_view', id=summary_id)
        else:
            # Người dùng đã hủy thao tác
            messages.info(request, 'Đã hủy thao tác chuyển tính lương')
            return redirect('attendance:attendance_summary_view', id=summary_id)

    return render(request, 'attendance/transfer_to_payroll.html', context)

@login_required
def daily_attendance_form(request):
    import logging
    logger = logging.getLogger(__name__)

    # GET request: Hiển thị form
    selected_date = request.GET.get('selected_date', datetime.now().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except ValueError:
        selected_date = datetime.now().date()

    if selected_date.weekday() == 6:
        messages.warning(request, "Không thể chấm công vào ngày Chủ nhật vì đây là ngày nghỉ!")
        employee_forms = []
    else:
        employees = Employee.objects.filter(
            models.Q(workshift__apply_to_all=True) | models.Q(workshift__employees__isnull=False)
        ).distinct()

        employee_forms = []
        for employee in employees:
            daily_attendance = DailyAttendance.objects.filter(
                employee=employee,
                date=selected_date
            ).first()
            initial_data = {
                'paid_work_days': daily_attendance.paid_work_days if daily_attendance else 1,
                'actual_work_days': daily_attendance.actual_work_days if daily_attendance else 1,
                'check_in_time': daily_attendance.check_in_time.strftime(
                    '%H:%M') if daily_attendance and daily_attendance.check_in_time else '',
                'check_out_time': daily_attendance.check_out_time.strftime(
                    '%H:%M') if daily_attendance and daily_attendance.check_out_time else '',
                'attendance_status': daily_attendance.attendance_status if daily_attendance else 'not_absent',
            }
            form = DailyAttendanceForm(initial=initial_data, prefix=f"employee_{employee.id}")
            employee_forms.append({
                'employee': employee,
                'form': form,
                'check_in_time': initial_data['check_in_time'],
                'check_out_time': initial_data['check_out_time'],
                'attendance_status': initial_data['attendance_status'],
                'paid_work_days': initial_data['paid_work_days'],
                'actual_work_days': initial_data['actual_work_days'],
            })

    if request.method == "POST":
        logger.debug(f"Request POST data: {request.POST}")
        selected_date_str = request.POST.get('selected_date')
        if not selected_date_str:
            messages.error(request, "Ngày không được để trống!")
            return redirect('attendance:daily_attendance_form')

        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Ngày không hợp lệ! Vui lòng chọn lại.")
            return redirect('attendance:daily_attendance_form')

        if selected_date.weekday() == 6:
            messages.warning(request, "Không thể chấm công vào ngày Chủ nhật vì đây là ngày nghỉ!")
            return redirect('attendance:daily_attendance_form')

        # Lấy danh sách employee_id từ request.POST
        employee_ids = [key.split('_')[-1] for key in request.POST.keys() if key.startswith('employee_id_')]
        logger.debug(f"Employee IDs from POST: {employee_ids}")
        if not employee_ids:
            messages.warning(request, "Không có nhân viên nào được gửi lên để chấm công!")
            return redirect('attendance:daily_attendance_form')

        employees = Employee.objects.filter(id__in=employee_ids)
        logger.debug(f"Employees from POST: {[emp.full_name for emp in employees]}")

        if not employees:
            messages.warning(request, "Không có nhân viên nào được gửi lên để chấm công!")
            return redirect('attendance:daily_attendance_form')

        employee_positions = list(set(employee.position for employee in employees))
        logger.debug(f"Employee positions: {employee_positions}")

        # Tìm hoặc tạo AttendanceRecord cho từng position
        attendance_records = {}
        for position in employee_positions:
            # Tìm AttendanceRecord phù hợp với ngày và position
            attendance_record = AttendanceRecord.objects.filter(
                start_date__lte=selected_date,
                end_date__gte=selected_date,
                positions=position
            ).first()

            if not attendance_record:
                # Tạo AttendanceRecord mới nếu không tìm thấy
                start_date = selected_date.replace(day=1)
                last_day = (selected_date.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
                end_date = last_day if last_day > selected_date else selected_date

                attendance_record = AttendanceRecord.objects.create(
                    name=f"Bảng chấm công tháng {selected_date.month}/{selected_date.year} - {position.name}",
                    start_date=start_date,
                    end_date=end_date,
                    attendance_type='manual',
                    positions=position,
                    apply_to_all_shifts=True
                )
                messages.info(request,
                              f"Đã tạo bảng chấm công chi tiết mới cho tháng {selected_date.month}/{selected_date.year} và vị trí {position.name}!")

            attendance_records[position.id] = attendance_record

        for employee in employees:
            form_prefix = f"employee_{employee.id}"
            # Kiểm tra xem các trường bắt buộc có giá trị hợp lệ không
            attendance_status = request.POST.get(f'{form_prefix}-attendance_status', '').strip()
            paid_work_days = request.POST.get(f'{form_prefix}-paid_work_days', '').strip()
            actual_work_days = request.POST.get(f'{form_prefix}-actual_work_days', '').strip()

            logger.debug(f"Data for employee {employee.full_name} (ID: {employee.id}): "
                         f"attendance_status={attendance_status}, "
                         f"paid_work_days={paid_work_days}, "
                         f"actual_work_days={actual_work_days}")

            if not attendance_status or not paid_work_days or not actual_work_days:
                logger.warning(
                    f"Missing or empty required fields for employee {employee.full_name} (ID: {employee.id})")
                messages.warning(request,
                                 f"Thiếu dữ liệu cho nhân viên {employee.full_name}: Vui lòng điền đầy đủ thông tin!")
                continue

            form = DailyAttendanceForm(request.POST, prefix=form_prefix)

            if form.is_valid():
                # Lấy AttendanceRecord phù hợp với position của nhân viên
                attendance_record = attendance_records[employee.position.id]

                daily_attendance, created = DailyAttendance.objects.get_or_create(
                    attendance_record=attendance_record,
                    employee=employee,
                    date=selected_date,
                    defaults={'attendance_status': 'not_absent'}
                )

                daily_attendance.paid_work_days = form.cleaned_data['paid_work_days']
                daily_attendance.actual_work_days = form.cleaned_data['actual_work_days']
                daily_attendance.check_in_time = form.cleaned_data['check_in_time']
                daily_attendance.check_out_time = form.cleaned_data['check_out_time']
                daily_attendance.attendance_status = form.cleaned_data['attendance_status']

                work_shift = WorkShift.objects.filter(
                    models.Q(employees=employee) | models.Q(apply_to_all=True)
                ).first()
                if not work_shift:
                    logger.warning(f"No work shift found for employee {employee.full_name}")
                    messages.warning(request, f"Không tìm thấy ca làm việc phù hợp cho nhân viên {employee.full_name}!")
                    continue
                daily_attendance.work_shift = work_shift
                daily_attendance.save()
                logger.debug(f"Saved DailyAttendance for employee {employee.full_name} on {selected_date}")
            else:
                logger.warning(f"Form invalid for employee {employee.full_name}: {form.errors}")
                messages.warning(request, f"Form không hợp lệ cho nhân viên {employee.full_name}: {form.errors}")

        messages.success(request, f"Đã chấm công cho ngày {selected_date} thành công!")
        return redirect('attendance:daily_attendance_form')

    context = {
        'selected_date': selected_date,
        'employee_forms': employee_forms,
    }
    return render(request, "attendance/daily_attendance_form.html", context)



@login_required
def attendance_dashboard(request):
    return render(request, 'attendance/dashboard.html')
